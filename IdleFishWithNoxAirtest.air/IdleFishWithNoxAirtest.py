# -*- encoding=utf8 -*-
__author__ = "tinyboxxx"


import openpyxl
import sys
import csv
from airtest.core.api import *
from poco.drivers.android.uiautomation import AndroidUiautomationPoco


#爬取数量
HowManyItemsToLog=100
#搜索名字
SearchItemName = 'iPad pro 10.5'
#最低最高价格筛选
LowPrice=1800
HighPrice=2500

#poco初始化
poco = AndroidUiautomationPoco(
    use_airtest_input=True, screenshot_each_action=False)
auto_setup(__file__)
#新建表格
workbook=openpyxl.Workbook()
worksheet = workbook.active
worksheet.cell(1,1,"Title")
worksheet.cell(1,2,"Price")
worksheet.cell(1,3,"想要")
rowsnow=2

inAppSearchNow
inAppStartSearch(SearchItemName,LowPrice,HighPrice)

#进入列表之后
items = poco(
    'com.taobao.idlefish:id/list_recyclerview').child('com.taobao.idlefish:id/card_root')

while(worksheet.max_row<HowManyItemsToLog):
    for i in range(1,4):
        itemNameString = items[i].offspring('com.taobao.idlefish:id/title_img').get_text().replace("\n", "").replace(",", " ").replace("  ", " ")
        itemPriceString = items[i].offspring('com.taobao.idlefish:id/integer_price').get_text()
        try:
            itemBuyerString = items[i].offspring('com.taobao.idlefish:id/search_item_mid_view').offspring('com.taobao.idlefish:id/search_item_flowlayout').offspring('android.widget.TextView').get_text().replace("人想要", " ")
        except:
            itemBuyerString3='0'
        worksheet.cell(rowsnow,1,itemNameString)
        worksheet.cell(rowsnow,2,itemPriceString)
        worksheet.cell(rowsnow,3,itemBuyerString)
        workbook.save(SearchItemName+'.xlsx') 
        rowsnow=worksheet.max_row+1


def controlFromLauncher(self):
    # 进入搜索界面
    poco("com.taobao.idlefish:id/search_bg_img_front").click()
    # 搜索框内输入文本
    poco("com.taobao.idlefish:id/search_term").click()
    
def inAppStartSearch(SearchItemName,LowPrice,HighPrice):
    #定义价格范围
    LowPrice=LowPrice
    HighPrice=HighPrice
    text( SearchItemName )
    #点击搜索框
    poco("com.taobao.idlefish:id/search_button").click()
    poco("com.taobao.idlefish:id/switch_search").click()
    #点击筛选
    touch(Template(r"tpl1583561091346.png", record_pos=(
        0.396, -0.699), resolution=(720, 1280)))
    #输入低价
    poco("com.taobao.idlefish:id/input_low_price").click()
    text( LowPrice )
    #输入高价
    poco("com.taobao.idlefish:id/input_high_price").click()
    text( HighPrice )
    #点击确定
    poco("com.taobao.idlefish:id/confirm").click()
    # 等待商品列表完全出现
    poco('com.taobao.idlefish:id/list_recyclerview').wait_for_appearance()

def swipeScreenNow():
    swipe(v1=[420, 1232], v2=[415, 330], duration=3)
    sleep(2.0)